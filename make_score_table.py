import sys
sys.stdout.reconfigure(encoding='utf-8')
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open('results/7fd7ae0e-529c-436a-a1ef-e479d7ef8b93/summary.json', encoding='utf-8') as f:
    data = json.load(f)
toda = next(p for p in data if '戸田' in p.get('name', ''))

scene_order = {'シーン1': 0, 'シーン2': 1, 'シーン3': 2}
answers = sorted(toda['answers'], key=lambda a: (scene_order.get(a['question_genre'], 9), a['question_id']))

COMPS = ['コミュニケーション', '情報収集', '情報分析', '想像・先読み', '計画']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '採点結果'

header_fill  = PatternFill('solid', start_color='1A3A5C')
scene_fills  = {
    'シーン1': PatternFill('solid', start_color='DDE8F0'),
    'シーン2': PatternFill('solid', start_color='E8F4EA'),
    'シーン3': PatternFill('solid', start_color='FEF3E2'),
}
thin   = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')

# タイトル
ws.merge_cells('A1:H1')
title = f'採点結果  {toda["name"]}  {toda["department"]}  {toda["total_score"]}/{toda["total_max"]}点（{toda["percentage"]}%）'
c = ws['A1']
c.value = title
c.font = Font(name='Arial', bold=True, size=12, color='1A3A5C')
c.alignment = center

# ヘッダー行
headers = ['シーン', '問題番号', 'コミュニケーション', '情報収集', '情報分析', '想像・先読み', '計画', '小計']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = border

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 10
for col in ['C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col].width = 16
ws.column_dimensions['H'].width = 8

# データ行
for i, ans in enumerate(answers, 1):
    row    = i + 2
    scene  = ans['question_genre']
    scores = ans['competency_scores']
    comps  = ans['competencies']
    fill   = scene_fills.get(scene, PatternFill())

    c1 = ws.cell(row=row, column=1, value=scene)
    c1.alignment = center
    c1.fill = fill
    c1.border = border
    c1.font = Font(name='Arial', size=10)

    c2 = ws.cell(row=row, column=2, value=f'Q{i:02d}')
    c2.alignment = center
    c2.fill = fill
    c2.border = border
    c2.font = Font(name='Arial', size=10)

    subtotal = 0
    for col, comp in enumerate(COMPS, 3):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.alignment = center
        cell.border    = border
        if comp in comps:
            score        = scores.get(comp, 0)
            cell.value   = score
            cell.font    = Font(name='Arial', size=10, bold=True)
            subtotal    += score
        else:
            cell.value = '-'
            cell.font  = Font(name='Arial', size=10, color='AAAAAA')

    c8 = ws.cell(row=row, column=8, value=subtotal)
    c8.alignment = center
    c8.fill   = fill
    c8.border = border
    c8.font   = Font(name='Arial', bold=True, size=10)

# 合計行
total_row = len(answers) + 3
ws.merge_cells(f'A{total_row}:B{total_row}')
ca = ws.cell(row=total_row, column=1, value='合計')
ca.font      = Font(name='Arial', bold=True, color='FFFFFF')
ca.fill      = header_fill
ca.alignment = center
ca.border    = border

for col, comp in enumerate(COMPS, 3):
    total = sum(ans['competency_scores'].get(comp, 0) for ans in answers if comp in ans['competencies'])
    max_v = sum(3 for ans in answers if comp in ans['competencies'])
    cell       = ws.cell(row=total_row, column=col, value=f'{total}/{max_v}')
    cell.font  = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill  = header_fill
    cell.alignment = center
    cell.border    = border

c8t = ws.cell(row=total_row, column=8, value=f'{toda["total_score"]}/{toda["total_max"]}')
c8t.font      = Font(name='Arial', bold=True, color='FFFFFF')
c8t.fill      = header_fill
c8t.alignment = center
c8t.border    = border

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 18

out = r'C:\Users\eshim\OneDrive\デスクトップ\戸田貴士_採点表.xlsx'
wb.save(out)
print('saved:', out)
